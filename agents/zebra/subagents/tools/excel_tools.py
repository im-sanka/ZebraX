"""
Excel Tools for Zebra Agent

Functions for reading, writing, and managing Excel files for data extraction.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional


def _normalize_boolean_value(value: Any) -> str:
    """
    Normalize boolean values to uppercase strings to prevent Excel auto-conversion.
    
    Args:
        value: Any value that might be a boolean
    
    Returns:
        "TRUE" or "FALSE" for boolean values, otherwise the original value as string
    """
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    
    # Check for string representations of booleans
    str_val = str(value).strip().lower()
    if str_val in ("true", "yes", "1"):
        return "TRUE"
    elif str_val in ("false", "no", "0"):
        return "FALSE"
    
    return str(value)


def excel_to_json(file_path: str) -> dict:
    """
    Reads an Excel file and converts its header to a JSON dictionary keys.
    The keys will be passed to LLM for data filling.

    Args:
        file_path: The path to the Excel file.
    
    Returns:
        Dictionary with column names as keys.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    df = pd.read_excel(file_path, nrows=0)
    json_template = {col: "" for col in df.columns}
    return json_template


def read_excel_data(file_path: str) -> List[Dict[str, Any]]:
    """
    Reads an Excel file and returns all rows as a list of dictionaries.

    Args:
        file_path: The path to the Excel file.
    
    Returns:
        List of dictionaries, each representing a row.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    df = pd.read_excel(file_path)
    # Replace NaN values with None for JSON serialization
    df = df.where(pd.notnull(df), None)
    # Convert to records and handle any remaining issues
    records = df.to_dict(orient='records')
    # Convert None to empty string for cleaner output
    for record in records:
        for key, value in record.items():
            if value is None or (isinstance(value, float) and pd.isna(value)):
                record[key] = ""
    return records


def update_excel_row(file_path: str, row_index: int, column_name: str, value: str) -> Dict[str, Any]:
    """
    Updates a specific cell in an Excel file.
    Boolean values (True/False) are automatically normalized to uppercase strings.

    Args:
        file_path: The path to the Excel file.
        row_index: The row index (0-based, excluding header).
        column_name: The column name to update.
        value: The value to set (as string).
    
    Returns:
        Dictionary with success status and message.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Use openpyxl to preserve string formatting
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find column index by name (row 1 = header)
        col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == column_name:
                col_idx = col
                break
        
        if col_idx is None:
            columns = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            return {"success": False, "error": f"Column '{column_name}' not found. Available: {columns}"}
        
        # row_index is 0-based (excluding header), Excel is 1-based with header at row 1
        excel_row = row_index + 2  # +1 for 1-based, +1 for header
        
        if excel_row > ws.max_row:
            return {"success": False, "error": f"Row index {row_index} out of range (0-{ws.max_row - 2})"}
        
        # Normalize boolean values to uppercase strings
        normalized_value = _normalize_boolean_value(value)
        
        ws.cell(row=excel_row, column=col_idx).value = normalized_value
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Updated row {row_index}, column '{column_name}' to '{normalized_value}'",
            "file_path": file_path
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def _batch_update_excel(file_path: str, updates: list) -> Dict[str, Any]:
    """
    INTERNAL FUNCTION - NOT EXPOSED AS A TOOL.
    Updates multiple cells in an Excel file in a single operation.
    If a column does not exist, it will be created automatically.

    Args:
        file_path: The path to the Excel file.
        updates: List of updates, each with 'row_index', 'column_name', and 'value'.
    
    Returns:
        Dictionary with success status and details.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        results = []
        columns_added = []
        
        for update in updates:
            row_index = update.get('row_index')
            column_name = update.get('column_name')
            value = update.get('value')
            
            # Auto-create column if it does not exist
            if column_name not in df.columns:
                df[column_name] = ""
                if column_name not in columns_added:
                    columns_added.append(column_name)
            
            if row_index < 0 or row_index >= len(df):
                results.append({"row": row_index, "column": column_name, "success": False, "error": "Row out of range"})
                continue
            
            df.at[row_index, column_name] = value
            results.append({"row": row_index, "column": column_name, "success": True, "value": value})
        
        df.to_excel(file_path, index=False)
        
        return {
            "success": True,
            "message": f"Batch update completed. {len([r for r in results if r['success']])} successful, {len([r for r in results if not r['success']])} failed.",
            "file_path": file_path,
            "columns_added": columns_added,
            "results": results
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def add_column_to_excel(file_path: str, column_name: str, default_value: str = "") -> Dict[str, Any]:
    """
    Adds a new column to an Excel file if it does not exist.

    Args:
        file_path: The path to the Excel file.
        column_name: The name of the column to add.
        default_value: The default value for the new column (default: empty string).
    
    Returns:
        Dictionary with success status and message.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        
        if column_name in df.columns:
            return {
                "success": True,
                "message": f"Column '{column_name}' already exists in Excel file",
                "file_path": file_path,
                "column_existed": True
            }
        
        df[column_name] = default_value
        df.to_excel(file_path, index=False)
        
        return {
            "success": True,
            "message": f"Added new column '{column_name}' to Excel file",
            "file_path": file_path,
            "column_existed": False
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_excel_columns(file_path: str) -> Dict[str, Any]:
    """
    Gets the list of column names from an Excel file.

    Args:
        file_path: The path to the Excel file.
    
    Returns:
        Dictionary with column names and count.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path, nrows=0)
        columns = list(df.columns)
        
        return {
            "success": True,
            "file_path": file_path,
            "columns": columns,
            "column_count": len(columns)
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def delete_excel_column(file_path: str, column_name: str) -> Dict[str, Any]:
    """
    Deletes a column from an Excel file.

    Args:
        file_path: The path to the Excel file.
        column_name: The name of the column to delete.
    
    Returns:
        Dictionary with success status and message.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        
        if column_name not in df.columns:
            return {
                "success": False,
                "error": f"Column '{column_name}' not found in Excel file. Available columns: {list(df.columns)}"
            }
        
        df = df.drop(columns=[column_name])
        df.to_excel(file_path, index=False)
        
        return {
            "success": True,
            "message": f"Deleted column '{column_name}' from Excel file",
            "file_path": file_path,
            "remaining_columns": list(df.columns)
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def delete_excel_row(file_path: str, row_index: int) -> Dict[str, Any]:
    """
    Deletes a row from an Excel file.

    Args:
        file_path: The path to the Excel file.
        row_index: The row index to delete (0-based, excluding header).
    
    Returns:
        Dictionary with success status and message.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        
        if row_index < 0 or row_index >= len(df):
            return {
                "success": False,
                "error": f"Row index {row_index} out of range (0-{len(df)-1})"
            }
        
        df = df.drop(index=row_index).reset_index(drop=True)
        df.to_excel(file_path, index=False)
        
        return {
            "success": True,
            "message": f"Deleted row {row_index} from Excel file",
            "file_path": file_path,
            "remaining_rows": len(df)
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_cell_value(file_path: str, row_index: int, column_name: str) -> Dict[str, Any]:
    """
    Gets the value of a specific cell in an Excel file.

    Args:
        file_path: The path to the Excel file.
        row_index: The row index (0-based, excluding header).
        column_name: The column name.
    
    Returns:
        Dictionary with the cell value.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        
        if column_name not in df.columns:
            return {
                "success": False,
                "error": f"Column '{column_name}' not found. Available columns: {list(df.columns)}"
            }
        
        if row_index < 0 or row_index >= len(df):
            return {
                "success": False,
                "error": f"Row index {row_index} out of range (0-{len(df)-1})"
            }
        
        value = df.at[row_index, column_name]
        # Handle NaN values
        if pd.isna(value):
            value = None
        
        return {
            "success": True,
            "row_index": row_index,
            "column_name": column_name,
            "value": value
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def clear_cell(file_path: str, row_index: int, column_name: str) -> Dict[str, Any]:
    """
    Clears (sets to empty) a specific cell in an Excel file.

    Args:
        file_path: The path to the Excel file.
        row_index: The row index (0-based, excluding header).
        column_name: The column name.
    
    Returns:
        Dictionary with success status and message.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        
        if column_name not in df.columns:
            return {
                "success": False,
                "error": f"Column '{column_name}' not found. Available columns: {list(df.columns)}"
            }
        
        if row_index < 0 or row_index >= len(df):
            return {
                "success": False,
                "error": f"Row index {row_index} out of range (0-{len(df)-1})"
            }
        
        old_value = df.at[row_index, column_name]
        df.at[row_index, column_name] = ""
        df.to_excel(file_path, index=False)
        
        return {
            "success": True,
            "message": f"Cleared cell at row {row_index}, column '{column_name}'",
            "old_value": old_value if not pd.isna(old_value) else None,
            "file_path": file_path
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_excel_info(file_path: str) -> Dict[str, Any]:
    """
    Gets general information about an Excel file (rows, columns, preview).

    Args:
        file_path: The path to the Excel file.
    
    Returns:
        Dictionary with file information.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        
        return {
            "success": True,
            "file_path": file_path,
            "row_count": len(df),
            "column_count": len(df.columns),
            "columns": list(df.columns),
            "first_5_rows": df.head().to_dict(orient='records')
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def data_template_filler(json_template: dict, paper_text: str, research_paper_title: str) -> dict:
    """
    Fills the provided JSON template with data based on the research paper content.

    For each key in the json_template, if the key is found in the paper text, 
    fill it with the research paper title.

    Args:
        json_template: The JSON template with placeholders.
        paper_text: The extracted text content of the research paper.
        research_paper_title: The title of the research paper to fill into the template.
    
    Returns:
        Filled template dictionary.
    """
    filled_template = json_template.copy()
    for key in filled_template:
        if key.lower() in paper_text.lower():
            if isinstance(filled_template[key], list):
                filled_template[key].append(research_paper_title)
            else:
                filled_template[key] = research_paper_title
    return filled_template


def list_pdf_files(directory: str) -> List[str]:
    """
    Lists all PDF files in a directory.

    Args:
        directory: The path to the directory.
    
    Returns:
        List of absolute paths to PDF files.
    """
    if not os.path.exists(directory):
        raise FileNotFoundError(f"Directory not found: {directory}")
    
    pdf_files = []
    for filename in os.listdir(directory):
        if filename.lower().endswith('.pdf'):
            pdf_files.append(os.path.join(directory, filename))
    
    return pdf_files


def batch_update_cells(file_path: str, updates: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Updates multiple cells in an Excel file in a single operation.
    Much more efficient than calling update_excel_row multiple times.
    Boolean values (True/False) are automatically normalized to uppercase strings.

    Args:
        file_path: The path to the Excel file.
        updates: List of updates, each dict with 'row_index', 'column_name', and 'value'.
                 Example: [{"row_index": 0, "column_name": "Status", "value": "TRUE"},
                          {"row_index": 1, "column_name": "Status", "value": "FALSE"}]
    
    Returns:
        Dictionary with success status and details.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Use openpyxl to preserve string formatting
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Build column name to index mapping
        col_map = {}
        for col in range(1, ws.max_column + 1):
            col_name = ws.cell(row=1, column=col).value
            if col_name:
                col_map[col_name] = col
        
        success_count = 0
        fail_count = 0
        columns_added = []
        
        for update in updates:
            row_index = update.get('row_index')
            column_name = update.get('column_name')
            value = update.get('value')
            
            # Auto-create column if it does not exist
            if column_name not in col_map:
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col).value = column_name
                col_map[column_name] = new_col
                columns_added.append(column_name)
            
            # row_index is 0-based (excluding header), Excel is 1-based with header at row 1
            excel_row = row_index + 2
            
            if excel_row > ws.max_row or row_index < 0:
                fail_count += 1
                continue
            
            # Normalize boolean values to uppercase strings
            normalized_value = _normalize_boolean_value(value)
            
            ws.cell(row=excel_row, column=col_map[column_name]).value = normalized_value
            success_count += 1
        
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Batch update completed: {success_count} cells updated, {fail_count} failed.",
            "file_path": file_path,
            "updated": success_count,
            "failed": fail_count,
            "columns_added": columns_added
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def transform_column(file_path: str, column_name: str, transformation: str) -> Dict[str, Any]:
    """
    Applies a transformation to ALL values in a column.
    Use this for bulk operations like capitalizing, uppercasing, etc.

    Args:
        file_path: The path to the Excel file.
        column_name: The name of the column to transform.
        transformation: The transformation to apply. Options:
                       - "uppercase": Convert to UPPERCASE (e.g., "true" becomes "TRUE")
                       - "lowercase": Convert to lowercase (e.g., "TRUE" becomes "true")
                       - "capitalize": Capitalize first letter (e.g., "true" becomes "True")
                       - "title": Title Case (e.g., "hello world" becomes "Hello World")
                       - "strip": Remove leading/trailing whitespace
    
    Returns:
        Dictionary with success status and details.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Use openpyxl directly to preserve string formatting (avoid Excel auto-converting TRUE/FALSE to booleans)
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find column index by name (row 1 = header)
        col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == column_name:
                col_idx = col
                break
        
        if col_idx is None:
            # Get all column names for error message
            columns = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            return {
                "success": False,
                "error": f"Column '{column_name}' not found. Available columns: {columns}"
            }
        
        transformation = transformation.lower().strip()
        valid_transformations = ["uppercase", "lowercase", "capitalize", "title", "strip"]
        if transformation not in valid_transformations:
            return {
                "success": False,
                "error": f"Unknown transformation '{transformation}'. Use: {', '.join(valid_transformations)}"
            }
        
        changed_count = 0
        total_rows = ws.max_row - 1  # Exclude header
        
        # Transform each cell in the column (starting from row 2 to skip header)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            original_value = str(cell.value) if cell.value is not None else ""
            
            if transformation == "uppercase":
                new_value = original_value.upper()
            elif transformation == "lowercase":
                new_value = original_value.lower()
            elif transformation == "capitalize":
                new_value = original_value.capitalize()
            elif transformation == "title":
                new_value = original_value.title()
            elif transformation == "strip":
                new_value = original_value.strip()
            
            if new_value != original_value:
                changed_count += 1
            
            # Write as explicit string to prevent Excel from converting to boolean
            cell.value = new_value
        
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Applied '{transformation}' to column '{column_name}'. {changed_count} values changed.",
            "file_path": file_path,
            "column": column_name,
            "transformation": transformation,
            "rows_affected": changed_count,
            "total_rows": total_rows
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def find_row_by_title(file_path: str, paper_title: str, title_column: str = "Title") -> Dict[str, Any]:
    """
    Finds the row index in an Excel file that matches a given paper title.
    Uses case-insensitive partial matching.

    Args:
        file_path: The path to the Excel file.
        paper_title: The title of the paper to find.
        title_column: The column name containing titles (default: "Title").
    
    Returns:
        Dictionary with row_index if found, or error if not found.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        df = pd.read_excel(file_path)
        
        if title_column not in df.columns:
            # Try to find a column that looks like a title column
            possible_title_cols = [col for col in df.columns if 'title' in col.lower()]
            if possible_title_cols:
                title_column = possible_title_cols[0]
            else:
                return {
                    "success": False,
                    "error": f"Title column '{title_column}' not found. Available columns: {list(df.columns)}"
                }
        
        paper_title_lower = paper_title.lower().strip()
        
        # Try exact match first
        for idx, row in df.iterrows():
            cell_value = str(row[title_column]).lower().strip() if pd.notna(row[title_column]) else ""
            if cell_value == paper_title_lower:
                return {
                    "success": True,
                    "row_index": int(idx),
                    "matched_title": row[title_column],
                    "match_type": "exact"
                }
        
        # Try partial match (paper title is contained in cell or vice versa)
        for idx, row in df.iterrows():
            cell_value = str(row[title_column]).lower().strip() if pd.notna(row[title_column]) else ""
            if paper_title_lower in cell_value or cell_value in paper_title_lower:
                return {
                    "success": True,
                    "row_index": int(idx),
                    "matched_title": row[title_column],
                    "match_type": "partial"
                }
        
        # Try word-based matching (at least 70% of words match)
        paper_words = set(paper_title_lower.split())
        best_match = None
        best_score = 0
        
        for idx, row in df.iterrows():
            cell_value = str(row[title_column]).lower().strip() if pd.notna(row[title_column]) else ""
            cell_words = set(cell_value.split())
            
            if not cell_words:
                continue
            
            # Calculate overlap score
            overlap = len(paper_words & cell_words)
            score = overlap / max(len(paper_words), len(cell_words))
            
            if score > best_score and score >= 0.5:  # At least 50% word overlap
                best_score = score
                best_match = (idx, row[title_column])
        
        if best_match:
            return {
                "success": True,
                "row_index": int(best_match[0]),
                "matched_title": best_match[1],
                "match_type": "fuzzy",
                "match_score": round(best_score, 2)
            }
        
        return {
            "success": False,
            "error": f"No matching row found for paper title: '{paper_title}'"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def update_classification_by_title(
    file_path: str,
    classifications: List[Dict[str, Any]],
    column_name: str,
    title_column: str = "Title"
) -> Dict[str, Any]:
    """
    Updates classification results in Excel by matching paper titles to rows.
    This is the PREFERRED method for saving classification results as it ensures
    values are placed in the correct rows by matching paper titles.

    Args:
        file_path: The path to the Excel file.
        classifications: List of classification results, each containing:
                        - 'title': The paper title to match
                        - 'result': The classification value (True/False or any string)
                        Optional:
                        - 'file': The PDF filename (used as fallback for title matching)
        column_name: The name of the column to update/create with classification values.
        title_column: The column in Excel containing paper titles (default: "Title").
    
    Returns:
        Dictionary with success status, matched/unmatched counts, and details.
    
    Example:
        update_classification_by_title(
            "test/table_1.xlsx",
            [{"title": "Paper A", "result": True}, {"title": "Paper B", "result": False}],
            "Regression Testing"
        )
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Load workbook for editing
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Read data for title matching
        df = pd.read_excel(file_path)
        
        # Find or determine title column
        if title_column not in df.columns:
            possible_title_cols = [col for col in df.columns if 'title' in col.lower()]
            if possible_title_cols:
                title_column = possible_title_cols[0]
            else:
                return {
                    "success": False,
                    "error": f"Title column '{title_column}' not found. Available columns: {list(df.columns)}"
                }
        
        # Build column name to index mapping
        col_map = {}
        for col in range(1, ws.max_column + 1):
            col_name = ws.cell(row=1, column=col).value
            if col_name:
                col_map[col_name] = col
        
        # Add column if it doesn't exist
        column_added = False
        if column_name not in col_map:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = column_name
            col_map[column_name] = new_col
            column_added = True
        
        matched = []
        unmatched = []
        
        for classification in classifications:
            paper_title = classification.get('title', '')
            result = classification.get('result', '')
            pdf_file = classification.get('file', '')
            
            # Normalize the result value
            normalized_value = _normalize_boolean_value(result)
            
            # Try to find matching row
            match_result = find_row_by_title(file_path, paper_title, title_column)
            
            if match_result['success']:
                row_index = match_result['row_index']
                excel_row = row_index + 2  # +1 for 1-based, +1 for header
                
                ws.cell(row=excel_row, column=col_map[column_name]).value = normalized_value
                matched.append({
                    "paper_title": paper_title,
                    "row_index": row_index,
                    "matched_title": match_result.get('matched_title'),
                    "match_type": match_result.get('match_type'),
                    "value": normalized_value
                })
            else:
                unmatched.append({
                    "paper_title": paper_title,
                    "pdf_file": pdf_file,
                    "error": match_result.get('error')
                })
        
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Classification update completed: {len(matched)} matched, {len(unmatched)} unmatched.",
            "file_path": file_path,
            "column_name": column_name,
            "column_added": column_added,
            "matched_count": len(matched),
            "unmatched_count": len(unmatched),
            "matched": matched,
            "unmatched": unmatched
        }
    except Exception as e:
        return {"success": False, "error": str(e)}
